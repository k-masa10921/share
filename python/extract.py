import openpyxl
print("sheetname??")
wb_r = openpyxl.load_workbook("サイトマップ.xlsx")
print(wb_r.sheetnames)
print("sheetname??")
name_s = input()
sheet_active = wb_r[name_s]
for row in sheet_active.iter_rows(min_row=4, min_col=2):
    addrs = []
    for cell in row:
        addrs.append(cell.coordinate)
    print(",".join(addrs))
