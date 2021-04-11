import openpyxl
import requests
print("username??")
user = input()
path = "C:/Users/"+user+"/Desktop/"
path_w = "C:/Users/"+user+"/Desktop/"
print("filename??")
name = input()
wb_r = openpyxl.load_workbook(path+name+".xlsx")
print(wb_r.sheetnames)

wb_w = openpyxl.Workbook()
default_ws = wb_w.worksheets[0]
default_ws.title = "200"
sheet_w = wb_w.create_sheet(title='not found')
sheet_w = wb_w.create_sheet(title='timeout')
sheet_w = wb_w.create_sheet(title='error')
for sheet_r in wb_r:
    for row in sheet_r.rows:
        addrs = []
        for cell_r in row:
            addrs.append(cell_r.value)
            for link in addrs:
                print(link)
                try:
                    resp = requests.get(link, timeout=(15))
                    print(resp.status_code)
                    if(resp.status_code == 200):
                        if(wb_r.index(sheet_r) == 0):
                            sheet_active = wb_w["200"]
                            sheet_active.append([link])
                    elif(resp.status_code == 404):
                        sheet_active = wb_w["not found"]
                        sheet_active.append([link])
                    else:
                        print("error")
                        sheet_active = wb_w["error"]
                        sheet_active.append([link])
                except requests.exceptions.Timeout as error:
                    print("timeout")
                    sheet_active = wb_w["timeout"]
                    sheet_active.append([link])
                    continue
wb_w.save(path_w+name+"_result.xlsx")
wb_w.close()
