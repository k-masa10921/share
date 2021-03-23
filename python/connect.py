import os
import glob
from pathlib import Path

import openpyxl
import csv

p = input()
name = input()
csvfiles = glob.glob(p+"/*.csv", recursive=False)
wb = openpyxl.Workbook()
for file in csvfiles:
    wb.create_sheet(os.path.splitext(os.path.basename(file))[0])
    wb.active = wb.sheetnames.index(
        os.path.splitext(os.path.basename(file))[0])
    ws = wb.active
    with open(file, encoding="shift-jis") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
wb.save(p+"/"+name+"_xml用URL.xlsx")
