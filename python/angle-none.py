import requests
from requests.auth import HTTPDigestAuth
import matplotlib.pyplot as plt
import io
from PIL import Image
import numpy as np
import cv2
import math
import openpyxl as excel

import numpy as np
import cv2
import glob

#カメラの高さHを設定cm
H = 17.8 + 70

#測定対象の高さhを設定cm
h = 12.84 + 5 + 1.5

s = H - h


f = 1.95

#カメラのIPアドレス　No.1 192.168.11.29:80, No.2 192.168.11.11
# 画像データの取得　　/cgi-bin/camera
# 解像度の指定　　　　resolution
url1 = "http://192.168.11.25:80/cgi-bin/camera"
url2 = "http://192.168.11.11/cgi-bin/camera"

# 認証情報
user1 = "tk1"
pswd1 = "takato01"

user2 = "tk2"
pswd2 = "takato02"

#閾値設定
threshold = 240

i = 1

wb = excel.Workbook()

while True:
    
    sheet = wb.active
    # 画像の取得
    rs1 = requests.get(url1, auth=HTTPDigestAuth(user1, pswd1))

    # 取得した画像データをOpenCVで扱う形式に変換1
    img_bin1 = io.BytesIO(rs1.content)   
    img_pil1 = Image.open(img_bin1)
    img_np1  = np.asarray(img_pil1)
    
    #グレースケール変換
    img1  = cv2.cvtColor(img_np1, cv2.COLOR_RGBA2GRAY)
    #２値化処理
    judge, img_thresh = cv2.threshold(img1, threshold, 255, cv2.THRESH_BINARY)
    #カーネルサイズの設定
    kernel = np.ones((5, 5), np.uint8)   
    #オープニング処理
    img_opening = cv2.morphologyEx(img_thresh, cv2.MORPH_OPEN, kernel)
    imgs = cv2.hconcat([img1, img_thresh])
    
    #重心の測定
    contours, _ = cv2.findContours(img_opening, cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    
    for c in contours:
        M = cv2.moments(c)
        #重心の位置計算(画像の左上基準　↓→)
        x = int(M["m10"] / M["m00"])
        y = int(M["m01"] / M["m00"])
        #重心の座標表示(画像の中心から　↑→)
        #(xの最大値=640,yの最大値=480)(単位はpx)
        x_grav = x - 320
        y_grav = -y + 240
        
        #重心の単位をmmに変換
        x_grav_mm = (x_grav / 320) * 1.76
        y_grav_mm = (y_grav / 240) * 1.32
        
        #カメラ座標系での位置をワールド座標系に変換　（－にする必要あり）
        Xc = (x_grav_mm * s) / y_grav_mm
        Yc = H - h
        Zc = (f * s) / y_grav_mm
        
        Xw = - Xc
        Yw = - Zc
        Zw = H - Yc
        print(Xw, Yw, Zw)
        
        
        sheet.cell(row=i,column=1).value = Xw
        sheet.cell(row=i,column=3).value = Yw
        wb.save("hello.xlsx")
        i = i + 1
    cv2.waitKey(1)
    

        
    #画像表示
    cv2.imshow("image1", imgs)
    cv2.waitKey(500)
cv2.destroyAllWindows()




